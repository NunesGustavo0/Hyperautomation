"""Gera o relatório executivo de conferência de lotes (RN01 a RN12).

Uso:
    python gerar_relatorio.py caminho/inspecao_lotes_10dias.xlsx

Sem argumento, o script procura o arquivo em ``data/input`` e em Downloads.
"""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import datetime
from pathlib import Path
from time import sleep
from typing import Callable, Collection

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


from src.validacao_lotes import (
    CLASSIFICACOES,
    RegistroValidado,
    texto
)
from src.item_processor import processar_item
from src.classificador_divergencia import (
    ClassificadorDivergencia,
)
from src.operational_indicators import (
    OperationalIndicators,
    consolidar_indicadores,
)
from src.ml_decisions import (
    AuditoriaDecisoesML,
    DecisaoML,
)
from src.auditoria_hibrida import AuditoriaPipelineHibrido
from src.base_referencia import ConfiguracaoRetryBase, consultar_base_com_retry

from src.dead_letter import RepositorioDeadLetter,processar_lote_com_dead_letter


CORES = {
    "Válido": "22C55E",
    "Divergência": "F59E0B",
    "Ambíguo": "8B5CF6",
    "Erro de Entrada": "EF4444",
}
ABAS = {
    "Válido": "Válidos",
    "Divergência": "Divergências",
    "Ambíguo": "Ambíguos",
    "Erro de Entrada": "Erros de Entrada",
}
ABAS_DIARIAS_ESPERADAS = tuple(
    f"Insp_{dia:02d}_06_2026" for dia in (15, 16, 17, 18, 19, 22, 23, 24, 25, 26)
)
TOTAIS_GABARITO = {
    "Válido": 150,
    "Divergência": 50,
    "Ambíguo": 20,
    "Erro de Entrada": 30,
}


def ler_e_validar(
    caminho: Path,
    auditoria_ml: AuditoriaPipelineHibrido | None = None,
    classificador: ClassificadorDivergencia | None = None,
    consulta_base_referencia: Callable[[], Collection[str]] | None = None,
    configuracao_retry_base: ConfiguracaoRetryBase | None = None,
    sleeper_base: Callable[[float], None] = sleep,
    repositorio_dead_letter: RepositorioDeadLetter | None = None,
    execution_id: str = "exec-processamento-local",
    correlation_id: str = "",
    max_tentativas_dado: int = 3,
) -> list[RegistroValidado]:
    if classificador is None:
        classificador = (
            ClassificadorDivergencia.de_configuracao()
        )

    if auditoria_ml is None:
        auditoria_ml = AuditoriaPipelineHibrido()

    if repositorio_dead_letter is None:
        repositorio_dead_letter = RepositorioDeadLetter(
            Path("data/output/dead_letter.jsonl")
        )

    planilha = pd.ExcelFile(caminho)

    abas_diarias = sorted(
        (aba for aba in planilha.sheet_names if aba.startswith("Insp_")),
        key=lambda aba: datetime.strptime(aba, "Insp_%d_%m_%Y"),
    )
    faltantes = sorted(set(ABAS_DIARIAS_ESPERADAS) - set(abas_diarias))
    inesperadas = sorted(set(abas_diarias) - set(ABAS_DIARIAS_ESPERADAS))
    if faltantes or inesperadas:
        raise ValueError(
            "As abas diárias não correspondem aos 10 dias esperados. "
            f"Faltantes: {faltantes or 'nenhuma'}; "
            f"inesperadas: {inesperadas or 'nenhuma'}."
        )
    if "Base_Referencia" not in planilha.sheet_names:
        raise ValueError("A aba obrigatória Base_Referencia não foi encontrada.")

    if consulta_base_referencia is None:
        def consulta_base_referencia() -> set[str]:
            referencia = pd.read_excel(
                caminho,
                sheet_name="Base_Referencia",
                header=1,
            )
            return {
                texto(valor)
                for valor in referencia["lote_id"]
                if texto(valor)
            }

    resultado_base = consultar_base_com_retry(
        consulta_base_referencia,
        configuracao=configuracao_retry_base,
        sleeper=sleeper_base,
    )
    lotes_referencia = set(resultado_base.lotes)
    resultados: list[RegistroValidado] = []



    for aba in abas_diarias:
        dados = pd.read_excel(caminho, sheet_name=aba, header=2)
        colunas_registro = [
            'lote_id',
            'produto',
            'linha',
            'turno',
            'status',
            'responsavel',
            'data',
            'observacao'
        ]

        # Remove somente rodapés/linhas sem os oito campos do registro.
        dados = dados[
            dados[colunas_registro]
            .notna()
            .any(axis=1)
        ]
        dados = dados[~dados["lote_id"].fillna("").astype(str).str.startswith("Total de registros:")]
        data_referencia = datetime.strptime(aba, "Insp_%d_%m_%Y").strftime("%d/%m/%Y")

        # RN11 é deliberadamente reiniciada a cada aba/dia.
        totais = Counter(texto(valor) for valor in dados["lote_id"] if texto(valor))
        vistas: Counter[str] = Counter()
        for _, registro in dados.iterrows():
            lote = texto(registro.get("lote_id"))

            if lote and totais[lote] > 1:
                vistas[lote] += 1
                ocorrencia = vistas[lote]
            else:
                ocorrencia = 1
            if not resultado_base.sucesso:
                # A indisponibilidade da base crítica impede uma decisão de
                # negócio segura. O item segue para revisão sem consultar ML.
                resultados.append(
                    RegistroValidado(
                        data_referencia=data_referencia,
                        lote=lote,
                        produto=texto(registro.get("produto")),
                        linha=texto(registro.get("linha")),
                        turno=texto(registro.get("turno")),
                        status="PENDENTE_REVISAO",
                        responsavel=texto(registro.get("responsavel")),
                        data_inspecao=texto(registro.get("data")),
                        observacao=texto(registro.get("observacao")),
                        classificacao="PENDENTE_REVISAO",
                        motivo=(
                            "Base de referência indisponível após "
                            f"{resultado_base.tentativas} tentativa(s)"
                        ),
                        acao_recomendada="Encaminhar para revisão humana",
                    )
                )
            else:

                def processar_registro(
                        _item,
                        registro_atual=registro,
                        data_atual=data_referencia,
                        ocorrencia_atual=ocorrencia,
                ) -> RegistroValidado:
                    """
                    Processa o registro original do Pandas.

                    O parâmetro _item é exigido pelo mecanismo de
                    dead letter, mas o processamento mantém a Series
                    original para não perder tipos e valores.
                    """

                    return processar_item(
                        registro=registro_atual,
                        data_referencia=data_atual,
                        lotes_referencia=lotes_referencia,
                        ocorrencia_no_dia=ocorrencia_atual,
                        classificador=classificador,
                        auditoria_ml=auditoria_ml,
                    )

                resultado_item = (
                    processar_lote_com_dead_letter(
                        [registro.to_dict()],
                        processar_registro,
                        repositorio=repositorio_dead_letter,
                        execution_id=execution_id,
                        correlation_id=correlation_id,
                        max_tentativas_dado=max_tentativas_dado,
                    )
                )

                # Os registros processados precisam voltar para a
                # lista utilizada pelo Bot B e pelo relatório.
                resultados.extend(
                    resultado_item.processados
                )
    return resultados

def formatar_rastreabilidade(ws) -> None:
    """
    Aplica formatação às colunas do pipeline híbrido.

    A coluna de confiança é exibida como percentual.
    A origem da decisão recebe uma cor para facilitar a conferência.
    """

    cabecalhos = {
        celula.value: celula.column
        for celula in ws[1]
        if celula.value
    }

    coluna_confianca = cabecalhos.get("Confiança ML")
    coluna_origem = cabecalhos.get("Origem da Decisão")

    if coluna_confianca is not None:
        for linha in range(2, ws.max_row + 1):
            celula = ws.cell(
                row=linha,
                column=coluna_confianca,
            )

            # O valor continua sendo float, por exemplo 0.92.
            # O Excel apenas mostra esse número como 92.00%.
            if celula.value is not None:
                celula.number_format = "0.00%"

    if coluna_origem is not None:
        for linha in range(2, ws.max_row + 1):
            celula = ws.cell(
                row=linha,
                column=coluna_origem,
            )

            if celula.value == "ml":
                celula.fill = PatternFill(
                    "solid",
                    fgColor="DCFCE7",
                )
                celula.font = Font(
                    color="166534",
                    bold=True,
                )

            elif celula.value == "fallback":
                celula.fill = PatternFill(
                    "solid",
                    fgColor="FEF3C7",
                )
                celula.font = Font(
                    color="92400E",
                    bold=True,
                )

def estilizar_tabela(ws, nome_tabela: str) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row >= 2:
        tabela = Table(displayName=nome_tabela, ref=ws.dimensions)
        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
        )
        ws.add_table(tabela)
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="17365D")
        celula.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for coluna in ws.columns:
        valores = [len(str(c.value or "")) for c in coluna[:80]]
        largura = min(max(max(valores, default=8) + 2, 11), 48)
        ws.column_dimensions[get_column_letter(coluna[0].column)].width = largura
    for row in ws.iter_rows(min_row=2):
        for celula in row:
            celula.alignment = Alignment(vertical="top", wrap_text=True)


def montar_resumo(
    ws,
    df: pd.DataFrame,
    momento: datetime,
    indicadores: OperationalIndicators,
) -> None:
    azul, azul_claro, branco = "17365D", "D9EAF7", "FFFFFF"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:N2")
    ws["A1"] = "CONFERÊNCIA DE LOTES · PAINEL EXECUTIVO"
    ws["A1"].font = Font(size=22, bold=True, color=branco)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34
    ws["A3"] = "Período analisado"
    ws["B3"] = f"{df['Data de Referência'].min()} a {df['Data de Referência'].max()}"
    ws["F3"] = "Atualizado em"
    ws["G3"] = momento.strftime("%d/%m/%Y %H:%M:%S")

    contagens = pd.Series(
        {
            "Válido": indicadores.validos,
            "Divergência": indicadores.divergencias,
            "Ambíguo": indicadores.ambiguos,
            "Erro de Entrada": indicadores.erros_entrada,
        }
    )
    total = indicadores.total_registros
    cards = [("TOTAL", total, "17365D")] + [
        (nome.upper(), int(contagens[nome]), CORES[nome]) for nome in CLASSIFICACOES
    ]
    for indice, (titulo, valor, cor) in enumerate(cards):
        coluna = 1 + indice * 3
        ws.merge_cells(start_row=5, start_column=coluna, end_row=5, end_column=coluna + 1)
        ws.merge_cells(start_row=6, start_column=coluna, end_row=6, end_column=coluna + 1)
        ws.merge_cells(start_row=7, start_column=coluna, end_row=7, end_column=coluna + 1)
        topo = ws.cell(5, coluna, titulo)
        numero = ws.cell(6, coluna, valor)
        for row in range(5, 8):
            for col in range(coluna, coluna + 2):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=cor)
        topo.font = Font(bold=True, color=branco, size=10)
        numero.font = Font(bold=True, color=branco, size=23)
        topo.alignment = numero.alignment = Alignment(horizontal="center", vertical="center")
        if titulo != "TOTAL":
            ws.cell(7, coluna).value = int(valor) / total if total else 0
            ws.cell(7, coluna).number_format = "0.0%"
            ws.cell(7, coluna).font = Font(bold=True, color=branco)
            ws.cell(7, coluna).alignment = Alignment(horizontal="center")

    ws["A10"] = "Distribuição por classificação"
    ws["A10"].font = Font(size=14, bold=True, color=azul)
    ws["A11"], ws["B11"], ws["C11"] = "Classificação", "Quantidade", "%"
    for i, nome in enumerate(CLASSIFICACOES, start=12):
        ws.cell(i, 1, nome)
        ws.cell(i, 2, int(contagens[nome]))
        ws.cell(i, 3, int(contagens[nome]) / total if total else 0)
        ws.cell(i, 3).number_format = "0.0%"

    ws["A51"] = "DEZ INDICADORES OPERACIONAIS"
    ws["A51"].font = Font(size=14, bold=True, color=azul)
    ws["A52"], ws["B52"] = "Indicador", "Resultado"
    indicadores_dashboard = (
        ("1. Total de registros", indicadores.total_registros, "0"),
        ("2. Registros válidos", indicadores.validos, "0"),
        ("3. Divergências", indicadores.divergencias, "0"),
        ("4. Registros ambíguos", indicadores.ambiguos, "0"),
        ("5. Erros de entrada", indicadores.erros_entrada, "0"),
        ("6. Regra mais acionada", indicadores.regra_mais_acionada, "@"),
        ("7. Taxa de retrabalho", indicadores.taxa_retrabalho, "0.0%"),
        ("8. Taxa de revisão humana", indicadores.taxa_revisao_humana, "0.0%"),
        ("9. Taxa de qualidade da entrada", indicadores.taxa_qualidade_entrada, "0.0%"),
        ("10. Ganho estimado de tempo (horas)", indicadores.ganho_estimado_horas, "0.00"),
    )
    for row, (nome, valor, formato) in enumerate(indicadores_dashboard, start=53):
        ws.cell(row, 1, nome)
        ws.cell(row, 2, valor)
        ws.cell(row, 2).number_format = formato

    rosca = DoughnutChart()
    rosca.title = "Distribuição dos 250 registros"
    rosca.holeSize = 58
    rosca.height, rosca.width = 8.2, 12.5
    rosca.add_data(Reference(ws, min_col=2, min_row=11, max_row=15), titles_from_data=True)
    rosca.set_categories(Reference(ws, min_col=1, min_row=12, max_row=15))
    rosca.dataLabels = DataLabelList()
    rosca.dataLabels.showPercent = True
    rosca.dataLabels.showLeaderLines = True
    rosca.series[0].data_points = [DataPoint(idx=i, spPr=None) for i in range(4)]
    for ponto, nome in zip(rosca.series[0].data_points, CLASSIFICACOES):
        ponto.graphicalProperties.solidFill = CORES[nome]
    ws.add_chart(rosca, "E10")

    ws["A28"] = "Evolução diária dos registros que exigem ação"
    ws["A28"].font = Font(size=14, bold=True, color=azul)
    ws["A29"], ws["B29"], ws["C29"], ws["D29"] = (
        "Data", "Divergências", "Ambíguos", "Total de problemas"
    )
    diario = (
        df.assign(_n=1)
        .pivot_table(index="Data de Referência", columns="Classificação", values="_n", aggfunc="sum", fill_value=0)
        .reindex(columns=CLASSIFICACOES, fill_value=0)
    )
    diario.index = pd.to_datetime(diario.index, format="%d/%m/%Y")
    diario = diario.sort_index()
    for row, (data, valores) in enumerate(diario.iterrows(), start=30):
        ws.cell(row, 1, data.to_pydatetime())
        ws.cell(row, 1).number_format = "dd/mm/yyyy"
        ws.cell(row, 2, int(valores["Divergência"]))
        ws.cell(row, 3, int(valores["Ambíguo"]))
        ws.cell(row, 4, int(valores["Divergência"] + valores["Ambíguo"] + valores["Erro de Entrada"]))

    linha = LineChart()
    linha.title = "Evolução dos registros"
    linha.y_axis.title = "Quantidade"
    linha.x_axis.title = "Dia da inspeção"
    linha.style = 13
    linha.height, linha.width = 9, 22
    linha.add_data(Reference(ws, min_col=2, max_col=4, min_row=29, max_row=29 + len(diario)), titles_from_data=True)
    linha.set_categories(Reference(ws, min_col=1, min_row=30, max_row=29 + len(diario)))
    for serie, cor in zip(linha.series, (CORES["Divergência"], CORES["Ambíguo"], CORES["Erro de Entrada"])):
        serie.graphicalProperties.line.solidFill = cor
        serie.graphicalProperties.line.width = 28575
        serie.marker.symbol = "circle"
        serie.marker.size = 7
    ws.add_chart(linha, "E28")

    ws["A43"] = "LEITURA PARA DECISÃO"
    ws["A44"] = "Corrigir na origem"
    ws["B44"] = f"{contagens['Erro de Entrada']} registros"
    ws["A45"] = "Conciliar com base/processo"
    ws["B45"] = f"{contagens['Divergência']} registros"
    ws["A46"] = "Decisão humana"
    ws["B46"] = f"{contagens['Ambíguo']} registros"
    ws["A48"] = "Nota: duplicidades são avaliadas separadamente em cada dia; apenas a partir da 2ª ocorrência."
    ws.merge_cells("A48:N48")
    ws["A48"].fill = PatternFill("solid", fgColor=azul_claro)
    ws["A48"].font = Font(italic=True, color=azul)

    borda = Border(bottom=Side(style="thin", color="CBD5E1"))
    for linha_celulas in ws.iter_rows(min_row=11, max_row=15, min_col=1, max_col=3):
        for celula in linha_celulas:
            celula.border = borda
    for coluna, largura in {"A": 25, "B": 18, "C": 14, "D": 20, "E": 14, "F": 16, "G": 20, "H": 14, "I": 14, "J": 14, "K": 14, "L": 14, "M": 14, "N": 14}.items():
        ws.column_dimensions[coluna].width = largura
    ws.freeze_panes = "A4"
    ws.print_area = "A1:N62"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def montar_ranking(ws, indicadores: OperationalIndicators) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.append(("Código da Regra", "Descrição", "Quantidade", "% do Total de Acionamentos"))
    for regra in indicadores.ranking_regras:
        ws.append((regra.codigo, regra.descricao, regra.quantidade, regra.percentual))
        ws.cell(ws.max_row, 4).number_format = "0.0%"
    estilizar_tabela(ws, "TabelaRankingRegras")


def montar_dicionario(ws) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.append(("Termo", "Definição em linguagem de negócio"))
    termos = (
        ("Divergência", "Registro que não coincide com a referência ou com as condições esperadas do processo."),
        ("Ambíguo", "Registro que precisa de decisão humana porque não permite uma conclusão automática segura."),
        ("Erro de Entrada", "Informação ausente ou inválida que deve ser corrigida na origem."),
        ("RN11", "Controle que identifica repetições do mesmo lote dentro do mesmo dia, a partir da segunda ocorrência."),
        ("Taxa de Retrabalho", "Parcela dos registros que exige correção ou conciliação."),
        ("Taxa de Revisão Humana", "Parcela dos registros encaminhada para uma decisão de uma pessoa."),
        ("Taxa de Qualidade da Entrada", "Parcela dos registros recebida sem erros de preenchimento."),
        ("Ganho Estimado de Tempo", "Estimativa didática das horas poupadas com a conferência automática dos registros válidos."),
    )
    for termo in termos:
        ws.append(termo)
    estilizar_tabela(ws, "TabelaDicionario")


def gerar_excel(
    registros: list[RegistroValidado],
    saida: Path,
    momento: datetime,
    indicadores: OperationalIndicators | None = None,
    decisoes_ml: list[DecisaoML] | tuple[DecisaoML, ...] | None = None,
) -> pd.DataFrame:
    """Gera as nove abas usando decisões de ML previamente registradas."""
    indicadores = indicadores or consolidar_indicadores(registros)
    df = pd.DataFrame(
        [
            registro.to_dict()
            for registro in registros
        ]
    )
    if decisoes_ml:
        # Mantém compatibilidade com a auditoria antiga.
        df_decisoes = pd.DataFrame(
            [
                decisao.to_excel_dict()
                for decisao in decisoes_ml
            ],
            columns=(
                "Lote ID",
                "Classe Prevista",
                "Probabilidade",
                "Nível de Confiança",
                "Latência (ms)",
                "Registrado em (UTC)",
                "Versão do Modelo",
            ),
        )
    else:
        # No fluxo atual, as decisões híbridas já estão
        # armazenadas nos registros produzidos pelo Bot B.
        decisoes_dos_registros = []

        for registro in registros:
            origem = (
                    registro.origem_decisao
                    or ""
            ).strip().lower()

            if origem not in {
                "ml",
                "fallback",
            }:
                continue

            decisoes_dos_registros.append(
                {
                    "Lote ID": registro.lote,
                    "Causa Provável": (
                        registro.causa_provavel
                    ),
                    "Origem da Decisão": origem,
                    "Confiança ML": (
                        registro.confianca_ml
                    ),
                    "Motivo do Fallback": (
                            registro.motivo_fallback
                            or None
                    ),
                    "Versão do Modelo": (
                        registro.versao_modelo
                    ),
                }
            )

        df_decisoes = pd.DataFrame(
            decisoes_dos_registros,
            columns=(
                "Lote ID",
                "Causa Provável",
                "Origem da Decisão",
                "Confiança ML",
                "Motivo do Fallback",
                "Versão do Modelo",
            ),
        )


    saida.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="Resumo", index=False)
        df.to_excel(writer, sheet_name="Todos", index=False)
        for classificacao, aba in ABAS.items():
            df[df["Classificação"] == classificacao].to_excel(writer, sheet_name=aba, index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Ranking de Regras", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Dicionário", index=False)
        df_decisoes.to_excel(writer, sheet_name="Decisões de ML", index=False)

    wb = load_workbook(saida)
    montar_resumo(wb["Resumo"], df, momento, indicadores)
    for numero, aba in enumerate(
            ("Todos", *ABAS.values()),
            start=1,
    ):
        worksheet = wb[aba]

        estilizar_tabela(
            worksheet,
            f"Tabela{numero}",
        )

        formatar_rastreabilidade(worksheet)
    montar_ranking(wb["Ranking de Regras"], indicadores)
    montar_dicionario(wb["Dicionário"])
    estilizar_tabela(wb["Decisões de ML"], "TabelaDecisoesML")
    aba_decisoes = wb[
        "Decisões de ML"
    ]

    cabecalhos_decisoes = {
        celula.value: celula.column
        for celula in aba_decisoes[1]
        if celula.value
    }

    for nome_coluna in (
            "Probabilidade",
            "Confiança ML",
    ):
        numero_coluna = (
            cabecalhos_decisoes.get(
                nome_coluna
            )
        )

        if numero_coluna is None:
            continue

        for linha in range(
                2,
                aba_decisoes.max_row + 1,
        ):
            celula = aba_decisoes.cell(
                row=linha,
                column=numero_coluna,
            )

            if isinstance(
                    celula.value,
                    (int, float),
            ):
                celula.number_format = (
                    "0.00%"
                )
    wb.active = wb.sheetnames.index("Resumo")
    wb.calculation.fullCalcOnLoad = True
    wb.save(saida)
    return df


def gerar_resumo_executivo(
    indicadores: OperationalIndicators,
    saida: Path,
) -> None:
    """Gera a leitura executiva usando o mesmo objeto entregue ao Excel."""
    destaque = (
        f"{indicadores.regra_mais_acionada} — "
        f"{indicadores.descricao_regra_mais_acionada}"
    )
    saida.write_text(
        "\n".join(
            (
                "# Resumo Executivo",
                "",
                "## Visão Geral",
                f"Foram analisados **{indicadores.total_registros} registros**: "
                f"**{indicadores.validos} válidos**, **{indicadores.divergencias} divergências**, "
                f"**{indicadores.ambiguos} ambíguos** e **{indicadores.erros_entrada} erros de entrada**.",
                "",
                "## Indicadores Principais",
                f"- Taxa de retrabalho: **{indicadores.taxa_retrabalho:.1%}**.",
                f"- Taxa de revisão humana: **{indicadores.taxa_revisao_humana:.1%}**.",
                f"- Taxa de qualidade da entrada: **{indicadores.taxa_qualidade_entrada:.1%}**.",
                "",
                "## Destaque",
                f"A regra mais acionada foi **{destaque}**.",
                "",
                "## Ganho Estimado de Tempo",
                f"O ganho estimado é de **{indicadores.ganho_estimado_horas:.2f} horas**, considerando "
                f"**{indicadores.minutos_poupados_por_registro_valido} minutos poupados por registro válido**.",
                "",
                "## Observação",
                "O ganho de tempo é uma estimativa didática baseada na premissa declarada acima; não representa medição de produtividade real.",
                "",
            )
        ),
        encoding="utf-8",
    )


def gerar_pdf_resumo(df: pd.DataFrame, saida: Path) -> bool:
    """Cria uma réplica estática do painel para impressão/entrega."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np
        from matplotlib.patches import Patch
    except ImportError:
        return False
    contagens = df["Classificação"].value_counts().reindex(CLASSIFICACOES, fill_value=0)
    diario = df.groupby(["Data de Referência", "Classificação"]).size().unstack(fill_value=0).reindex(columns=CLASSIFICACOES, fill_value=0)
    diario.index = pd.to_datetime(diario.index, format="%d/%m/%Y")
    diario = diario.sort_index()
    fig = plt.figure(figsize=(16, 9), facecolor="#F7FAFC")
    grade = fig.add_gridspec(3, 5, height_ratios=[0.7, 2.7, 0.8], hspace=0.45, wspace=0.5)
    fig.suptitle("CONFERÊNCIA DE LOTES · PAINEL EXECUTIVO", fontsize=22, fontweight="bold", color="#17365D", y=0.97)
    cards = [("TOTAL", len(df), "#17365D")] + [(n.upper(), int(contagens[n]), "#" + CORES[n]) for n in CLASSIFICACOES]
    for i, (nome, valor, cor) in enumerate(cards):
        ax = fig.add_subplot(grade[0, i]); ax.axis("off")
        pct = "" if nome == "TOTAL" else f"\n{valor / len(df):.1%}"
        ax.text(0.5, 0.5, f"{nome}\n{valor}{pct}", ha="center", va="center", color="white", fontsize=12, fontweight="bold", bbox=dict(boxstyle="round,pad=0.8", facecolor=cor, edgecolor=cor))
    ax1 = fig.add_subplot(grade[1, :2])
    cores_distribuicao = ["#" + CORES[n] for n in CLASSIFICACOES]
    ax1.pie(
        contagens,
        labels=None,
        autopct="%1.1f%%",
        pctdistance=0.78,
        startangle=90,
        colors=cores_distribuicao,
        textprops={"color": "white", "fontweight": "bold", "fontsize": 11},
        wedgeprops=dict(width=0.48, edgecolor="white", linewidth=2),
    )
    ax1.set_title("Distribuição dos registros", fontweight="bold", color="#17365D")
    legenda = [
        Patch(facecolor=cor, edgecolor="none", label=nome)
        for nome, cor in zip(CLASSIFICACOES, cores_distribuicao)
    ]
    ax1.legend(
        handles=legenda,
        loc="lower left",
        bbox_to_anchor=(-0.12, -0.18),
        ncol=2,
        frameon=True,
        fancybox=True,
        framealpha=1,
        facecolor="white",
        edgecolor="#CBD5E1",
        fontsize=9,
        title="Legenda",
        title_fontproperties={"weight": "bold", "size": 9},
    )
    ax2 = fig.add_subplot(grade[1, 2:])
    eixo_x = diario.index.strftime("%d/%m")
    total_problemas = diario[["Divergência", "Ambíguo", "Erro de Entrada"]].sum(axis=1)
    posicoes = np.arange(len(eixo_x))
    largura = 0.25
    barras_divergencias = ax2.bar(
        posicoes - largura,
        diario["Divergência"],
        largura,
        label="Divergências",
        color="#F59E0B",
    )
    barras_ambiguos = ax2.bar(
        posicoes,
        diario["Ambíguo"],
        largura,
        label="Ambíguos",
        color="#8B5CF6",
    )
    barras_total = ax2.bar(
        posicoes + largura,
        total_problemas,
        largura,
        label="Total de problemas",
        color="#EF4444",
    )
    for barras in (barras_divergencias, barras_ambiguos, barras_total):
        ax2.bar_label(barras, padding=2, fontsize=7, color="#334155")
    ax2.set_title("Evolução dos registros", fontweight="bold", color="#17365D")
    ax2.set_ylabel("Quantidade")
    ax2.set_xticks(posicoes, eixo_x, rotation=35, ha="right")
    ax2.set_ylim(0, max(total_problemas) + 4)
    ax2.grid(axis="y", alpha=.25)
    ax2.set_axisbelow(True)
    ax2.legend(loc="upper left", ncol=3, frameon=False, fontsize=8)
    ax3 = fig.add_subplot(grade[2, :]); ax3.axis("off")
    ax3.text(0, .75, f"Corrigir na origem: {contagens['Erro de Entrada']}  |  Conciliar: {contagens['Divergência']}  |  Decisão humana: {contagens['Ambíguo']}", fontsize=13, fontweight="bold", color="#17365D")
    ax3.text(0, .25, "Duplicidades são avaliadas separadamente em cada dia e somente a partir da 2ª ocorrência.", fontsize=10, color="#475569")
    fig.savefig(saida, bbox_inches="tight")
    plt.close(fig)
    return True


def salvar_log(df: pd.DataFrame, caminho: Path, origem: Path, momento: datetime) -> None:
    contagens = df["Classificação"].value_counts().reindex(CLASSIFICACOES, fill_value=0)
    problemas = int(contagens[["Divergência", "Ambíguo", "Erro de Entrada"]].sum())
    resultado_obtido = {nome: int(contagens[nome]) for nome in CLASSIFICACOES}
    aceite = len(df) == 250 and resultado_obtido == TOTAIS_GABARITO
    linhas = [
        "LOG DE EXECUÇÃO — CONFERÊNCIA DE LOTES",
        f"Data/hora: {momento.strftime('%d/%m/%Y %H:%M:%S')}",
        f"Arquivo de origem: {origem.resolve()}",
        f"Total processado: {len(df)}",
        *(f"{nome}: {int(contagens[nome])}" for nome in CLASSIFICACOES),
        f"Total de registros problemáticos (todas as categorias): {problemas}",
        "Gabarito: 150 válidos + 50 divergências + 20 ambíguos + 30 erros de entrada = 250.",
        "RN11: contagem reiniciada em cada aba diária; divergência somente a partir da 2ª ocorrência.",
        "Validação de aceite: " + ("APROVADA" if aceite else "REVISAR"),
    ]
    # BOM facilita a abertura correta em Bloco de Notas/PowerShell legados.
    caminho.write_text("\n".join(linhas) + "\n", encoding="utf-8-sig")


def localizar_entrada(argumento: str | None) -> Path:
    candidatos = [
        Path(argumento) if argumento else None,
        Path("data/input/inspecao_lotes_10dias.xlsx"),
        Path.home() / "Downloads" / "inspecao_lotes_10dias.xlsx",
    ]
    for candidato in candidatos:
        if candidato and candidato.exists():
            return candidato
    raise FileNotFoundError("Informe o caminho de inspecao_lotes_10dias.xlsx.")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "entrada",
        nargs="?",
        help="Planilha de inspeções",
    )
    parser.add_argument(
        "--saida",
        default="reports/relatorio_conferencia_lotes.xlsx",
    )

    args = parser.parse_args()

    origem = localizar_entrada(args.entrada)
    saida = Path(args.saida)
    momento = datetime.now()

    auditoria_ml = AuditoriaPipelineHibrido()

    registros = ler_e_validar(
        origem,
        auditoria_ml=auditoria_ml,
    )

    indicadores = consolidar_indicadores(registros)

    df = gerar_excel(
        registros,
        saida,
        momento,
        indicadores,
    )

    gerar_resumo_executivo(
        indicadores,
        saida.with_name("resumo_executivo.md"),
    )

    salvar_log(
        df,
        saida.with_name("log_execucao.txt"),
        origem,
        momento,
    )

    pdf_ok = gerar_pdf_resumo(
        df,
        saida.with_name("dashboard_resumo.pdf"),
    )

    contagens = (
        df["Classificação"]
        .value_counts()
        .reindex(CLASSIFICACOES, fill_value=0)
    )

    print(f"Relatório: {saida.resolve()}")
    print(
        f"Total: {len(df)} | "
        + " | ".join(
            f"{nome}: {int(contagens[nome])}"
            for nome in CLASSIFICACOES
        )
    )

    print(
        "PDF: gerado"
        if pdf_ok
        else "PDF: matplotlib não instalado; Excel gerado normalmente"
    )


if __name__ == "__main__":
    main()
