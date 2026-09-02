"""Fluxo local completo dos seis bots e seus três simuladores."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from executar_pipeline_capstone import BOTS, executar_pipeline_capstone


def test_pipeline_capstone_completo_preserva_artefatos_ids_e_processos(tmp_path):
    resultado = executar_pipeline_capstone(
        diretorio_saida=tmp_path,
        alertas="console",
    )

    assert resultado.sucesso is True
    assert resultado.total_desktop == 12
    assert resultado.total_web == 12
    assert resultado.total_consolidado == 12
    assert resultado.processos_encerrados is True

    relatorio = Path(resultado.caminho_relatorio)
    workbook = load_workbook(relatorio, read_only=True)
    try:
        assert workbook["Consolidado"].max_row == 13
    finally:
        workbook.close()

    artefatos = (
        tmp_path / "estoque_desktop.json",
        tmp_path / "pedidos_fornecedores.json",
        tmp_path / "registros_consolidados.json",
        tmp_path / "registros_classificados.json",
        tmp_path / "resumo_execucao.json",
    )
    assert all(caminho.is_file() for caminho in artefatos)

    conteudo_logs = Path(resultado.caminho_logs).read_text(encoding="utf-8")
    eventos = [json.loads(linha) for linha in conteudo_logs.splitlines()]
    assert set(BOTS).issubset({evento.get("bot_id") for evento in eventos})
    assert {
        evento["execution_id"] for evento in eventos if "execution_id" in evento
    } == {resultado.execution_id}
    assert {
        evento["correlation_id"]
        for evento in eventos
        if "correlation_id" in evento
    } == {resultado.correlation_id}

    for caminho in artefatos:
        conteudo = caminho.read_text(encoding="utf-8")
        assert resultado.execution_id in conteudo
        assert resultado.correlation_id in conteudo
